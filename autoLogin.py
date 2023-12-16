# tips
# selenium==3.141.0
# urllib3==1.26.2

from selenium import webdriver
from PIL import Image
from selenium.webdriver.common.by import By
import openpyxl
import ddddocr
import time

type = 0
account_start = 0
acount_end = 0

wb = openpyxl.load_workbook("data.xlsx") # reading data from data.xlsx
s1 = wb["工作表1"] # reading chart which names 工作表1

account = []
password = []

# image recongnization
def ValidationCodeRec(name):
	with open(name, "rb") as f:
		img_bytes = f.read()
	
	result = ocr.classification(img_bytes)
	return result

# main part
max_row = s1.max_row # get max_row in the chart

print("\n\n")
print("type 0: auto vote for all account in the data chart")
print("type 1: vote for account in the data chart from a to b")
print("other type: exit")
print("\n\n")

type = int(input("choose type which you want to use:"))

if(type == 0):
	total_try = int(max_row * 1.5)

	# add data in chart to account & password
	for i in range(1, max_row + 1):
		account.append(s1.cell(i, 1).value)
		password.append(s1.cell(i, 2).value)
elif(type == 1):
	account_start = int(input("input a:"))
	acount_end = int(input("input b:"))

	total_try = int((acount_end - account_start) * 1.5)

	# add data in chart to account & password
	for i in range(account_start, acount_end + 1):
		account.append(s1.cell(i, 1).value)
		password.append(s1.cell(i, 2).value)
else:
	exit()



# image recongize part
ocr = ddddocr.DdddOcr()

times = 0

for i in account:
	# webpage par
	driver = webdriver.Chrome(executable_path="chromedriver/chromedriver")
	url = "https://youthdream.phdf.org.tw/member/login"
	url1 = "https://youthdream.phdf.org.tw/project/show?keyword=南友&page=1"
	driver.get(url)

	# maximu, the window
	driver.maximize_window()

	# input account & password
	driver.find_element_by_name("email").send_keys(account[times])
	driver.find_element_by_name("password").send_keys(password[times])

	now_try = 0

	# check is it successly login
	while(1):
		now_try = now_try + 1

		if(now_try > total_try): 
			print("it is a disaster")
			break

		try:
			user_mobile = driver.find_element_by_name("mobile")
		except:
			# log in failure -> try again
			print("failure")

			# save screen shot of the page
			driver.save_screenshot(str(times) + ".png")

			# get item of captcha code
			validation = driver.find_element_by_id("captcha")

			# get four point of the validation cade image
			left = validation.location["x"] + 150
			right = validation.location["x"] + validation.size["width"] + 300
			top = validation.location["y"] - 960
			bottom = validation.location["y"] + validation.size["height"] - 940

			# cut validation img out
			img = Image.open(str(times) + ".png")
			img = img.crop((left, top, right, bottom))
			img.save(str(times) + ".png")

			# get code by ddddocr
			code = ValidationCodeRec(str(times) + ".png")

			# input validation cade
			driver.find_element_by_name("captcha").clear()
			driver.find_element_by_name("captcha").send_keys(code)

			# click login
			driver.find_element(By.CSS_SELECTOR, ".button-5").click()

			# wait until the page loading done
			driver.implicitly_wait(2)
		else:
			print("success")

			driver.execute_script('window.open("https://youthdream.phdf.org.tw/project/show?keyword=南友&page=1")')

			# change to TSA page
			driver.switch_to_window(driver.window_handles[1])
			driver.find_element(By.CSS_SELECTOR, ".button-2.vote-project").click()	# click vote

			# wait until the page loading done
			driver.implicitly_wait(2)

			break

	times = times + 1
	time.sleep(3)
	driver.quit()